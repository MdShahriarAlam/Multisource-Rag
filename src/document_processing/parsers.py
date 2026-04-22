"""Document parsers for various file formats."""
import csv
import io
import json
from abc import ABC, abstractmethod
from typing import Dict, List, Type


class BaseParser(ABC):
    """Abstract base class for document parsers."""

    @abstractmethod
    def parse(self, content: bytes, filename: str) -> List[Dict[str, str]]:
        """Parse document content into text segments.

        Returns list of dicts with at least: text, metadata (page_number, etc.)
        """


class PDFParser(BaseParser):
    """Parse PDF files using pypdf."""

    def parse(self, content: bytes, filename: str) -> List[Dict[str, str]]:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(content))
        segments = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text and text.strip():
                segments.append(
                    {"text": text.strip(), "page_number": str(i + 1)}
                )
        return segments


class DocxParser(BaseParser):
    """Parse Word documents using python-docx."""

    def parse(self, content: bytes, filename: str) -> List[Dict[str, str]]:
        from docx import Document

        doc = Document(io.BytesIO(content))
        paragraphs = []
        current_text = []

        for para in doc.paragraphs:
            if para.text.strip():
                current_text.append(para.text.strip())
            elif current_text:
                paragraphs.append(
                    {"text": "\n".join(current_text), "section": str(len(paragraphs) + 1)}
                )
                current_text = []

        if current_text:
            paragraphs.append(
                {"text": "\n".join(current_text), "section": str(len(paragraphs) + 1)}
            )

        return paragraphs


class TextParser(BaseParser):
    """Parse plain text files."""

    def parse(self, content: bytes, filename: str) -> List[Dict[str, str]]:
        text = content.decode("utf-8", errors="replace")
        if text.strip():
            return [{"text": text.strip()}]
        return []


class CSVParser(BaseParser):
    """Parse CSV files — groups rows into text blocks with headers."""

    ROWS_PER_CHUNK = 25

    def parse(self, content: bytes, filename: str) -> List[Dict[str, str]]:
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))

        rows = list(reader)
        if not rows:
            return []

        header = rows[0]
        data_rows = rows[1:]
        segments = []

        for i in range(0, len(data_rows), self.ROWS_PER_CHUNK):
            chunk_rows = data_rows[i : i + self.ROWS_PER_CHUNK]
            lines = [", ".join(header)]
            for row in chunk_rows:
                lines.append(", ".join(row))
            segments.append(
                {
                    "text": "\n".join(lines),
                    "row_range": f"{i + 1}-{i + len(chunk_rows)}",
                }
            )

        return segments


class ExcelParser(BaseParser):
    """Parse Excel files using openpyxl."""

    ROWS_PER_CHUNK = 25

    def parse(self, content: bytes, filename: str) -> List[Dict[str, str]]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        segments = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = [str(c) if c is not None else "" for c in rows[0]]
            data_rows = rows[1:]

            for i in range(0, len(data_rows), self.ROWS_PER_CHUNK):
                chunk_rows = data_rows[i : i + self.ROWS_PER_CHUNK]
                lines = [", ".join(header)]
                for row in chunk_rows:
                    lines.append(
                        ", ".join(str(c) if c is not None else "" for c in row)
                    )
                segments.append(
                    {
                        "text": "\n".join(lines),
                        "sheet": sheet_name,
                        "row_range": f"{i + 1}-{i + len(chunk_rows)}",
                    }
                )

        wb.close()
        return segments


class JSONParser(BaseParser):
    """Parse JSON files — flattens into readable text."""

    def parse(self, content: bytes, filename: str) -> List[Dict[str, str]]:
        text = content.decode("utf-8", errors="replace")
        data = json.loads(text)

        if isinstance(data, list):
            # Array of objects — group into chunks
            segments = []
            chunk_size = 10
            for i in range(0, len(data), chunk_size):
                chunk = data[i : i + chunk_size]
                formatted = json.dumps(chunk, indent=2, default=str)
                segments.append(
                    {"text": formatted, "item_range": f"{i + 1}-{i + len(chunk)}"}
                )
            return segments
        else:
            # Single object
            formatted = json.dumps(data, indent=2, default=str)
            return [{"text": formatted}]


# Extension -> Parser mapping
_PARSER_MAP: Dict[str, Type[BaseParser]] = {
    ".pdf": PDFParser,
    ".docx": DocxParser,
    ".txt": TextParser,
    ".csv": CSVParser,
    ".xlsx": ExcelParser,
    ".json": JSONParser,
}


class ParserFactory:
    """Factory to get the right parser for a file extension."""

    @staticmethod
    def get_parser(filename: str) -> BaseParser:
        ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        parser_cls = _PARSER_MAP.get(ext)
        if parser_cls is None:
            raise ValueError(f"No parser available for extension '{ext}'")
        return parser_cls()

    @staticmethod
    def supported_extensions() -> List[str]:
        return list(_PARSER_MAP.keys())
